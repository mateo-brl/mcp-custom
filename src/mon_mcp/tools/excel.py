"""Outils MCP pour la manipulation de fichiers Excel et CSV."""

import csv
import io
import json
from datetime import datetime, date, time
from pathlib import Path

MAX_ROWS = 10000


def _convert_cell(value):
    """Convertit une valeur de cellule pour la serialisation JSON."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value


def lire_excel(chemin: str, feuille: str = "") -> str:
    """
    Lit un fichier Excel (.xlsx).

    Args:
        chemin: Chemin du fichier Excel
        feuille: Nom de la feuille (defaut: feuille active)

    Returns:
        Les donnees en JSON (colonnes + lignes).
    """
    try:
        import openpyxl
    except ImportError:
        return "Erreur: openpyxl non installe. pip install openpyxl"

    try:
        path = Path(chemin)
        if not path.exists():
            return f"Erreur: le fichier '{chemin}' n'existe pas"

        wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)

        if feuille:
            if feuille not in wb.sheetnames:
                wb.close()
                return f"Erreur: feuille '{feuille}' introuvable. Feuilles: {', '.join(wb.sheetnames)}"
            ws = wb[feuille]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return json.dumps({
                "fichier": str(path.resolve()),
                "feuille": ws.title,
                "colonnes": [],
                "lignes": 0,
                "donnees": [],
            }, ensure_ascii=False)

        colonnes = [str(c) if c is not None else f"Colonne_{i+1}" for i, c in enumerate(rows[0])]
        donnees = []
        tronque = False

        for row in rows[1:MAX_ROWS + 1]:
            donnees.append({
                col: _convert_cell(val)
                for col, val in zip(colonnes, row)
            })

        if len(rows) - 1 > MAX_ROWS:
            tronque = True

        result = {
            "fichier": str(path.resolve()),
            "feuille": ws.title,
            "colonnes": colonnes,
            "lignes": len(donnees),
            "donnees": donnees,
        }
        if tronque:
            result["tronque"] = True
            result["total_lignes"] = len(rows) - 1

        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return f"Erreur: {str(e)}"


def ecrire_excel(chemin: str, donnees: str, feuille: str = "Sheet1") -> str:
    """
    Ecrit ou cree un fichier Excel (.xlsx) a partir de donnees JSON.

    Args:
        chemin: Chemin du fichier Excel a creer/ecrire
        donnees: Donnees en JSON — liste de listes OU liste de dicts
        feuille: Nom de la feuille (defaut: "Sheet1")

    Returns:
        Confirmation avec le nombre de lignes ecrites.
    """
    try:
        import openpyxl
    except ImportError:
        return "Erreur: openpyxl non installe. pip install openpyxl"

    try:
        data = json.loads(donnees)
        if not isinstance(data, list) or len(data) == 0:
            return "Erreur: les donnees doivent etre une liste non vide"

        Path(chemin).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = feuille

        # Format: liste de dicts → extraire les cles comme en-tetes
        if isinstance(data[0], dict):
            colonnes = list(data[0].keys())
            ws.append(colonnes)
            for row in data:
                ws.append([row.get(col) for col in colonnes])
        # Format: liste de listes → premiere ligne = en-tetes
        elif isinstance(data[0], list):
            for row in data:
                ws.append(row)
        else:
            wb.close()
            return "Erreur: format non supporte. Utilisez une liste de dicts ou de listes."

        wb.save(chemin)
        wb.close()

        nb_lignes = len(data)
        return f"Fichier Excel ecrit: '{Path(chemin).resolve()}' ({nb_lignes} lignes, feuille '{feuille}')"
    except json.JSONDecodeError as e:
        return f"Erreur JSON: {str(e)}"
    except Exception as e:
        return f"Erreur: {str(e)}"


def lire_csv(chemin: str, separateur: str = ",", encodage: str = "utf-8") -> str:
    """
    Lit un fichier CSV.

    Args:
        chemin: Chemin du fichier CSV
        separateur: Separateur de colonnes (defaut: ",")
        encodage: Encodage du fichier (defaut: "utf-8")

    Returns:
        Les donnees en JSON.
    """
    try:
        path = Path(chemin)
        if not path.exists():
            return f"Erreur: le fichier '{chemin}' n'existe pas"

        with open(chemin, "r", encoding=encodage, newline="") as f:
            reader = csv.reader(f, delimiter=separateur)
            rows = []
            for i, row in enumerate(reader):
                if i > MAX_ROWS:
                    break
                rows.append(row)

        if not rows:
            return json.dumps({
                "fichier": str(path.resolve()),
                "colonnes": [],
                "lignes": 0,
                "donnees": [],
            }, ensure_ascii=False)

        colonnes = rows[0]
        donnees = []
        for row in rows[1:]:
            donnees.append({
                col: val for col, val in zip(colonnes, row)
            })

        result = {
            "fichier": str(path.resolve()),
            "separateur": separateur,
            "colonnes": colonnes,
            "lignes": len(donnees),
            "donnees": donnees,
        }
        if len(rows) - 1 > MAX_ROWS:
            result["tronque"] = True

        return json.dumps(result, ensure_ascii=False, indent=2)
    except UnicodeDecodeError:
        return f"Erreur: impossible de lire '{chemin}' avec l'encodage {encodage}"
    except Exception as e:
        return f"Erreur: {str(e)}"


def ecrire_csv(chemin: str, donnees: str, separateur: str = ",", encodage: str = "utf-8") -> str:
    """
    Ecrit un fichier CSV a partir de donnees JSON.

    Args:
        chemin: Chemin du fichier CSV a creer/ecrire
        donnees: Donnees en JSON — liste de listes OU liste de dicts
        separateur: Separateur de colonnes (defaut: ",")
        encodage: Encodage du fichier (defaut: "utf-8")

    Returns:
        Confirmation avec le nombre de lignes ecrites.
    """
    try:
        data = json.loads(donnees)
        if not isinstance(data, list) or len(data) == 0:
            return "Erreur: les donnees doivent etre une liste non vide"

        Path(chemin).parent.mkdir(parents=True, exist_ok=True)

        with open(chemin, "w", encoding=encodage, newline="") as f:
            writer = csv.writer(f, delimiter=separateur)

            if isinstance(data[0], dict):
                colonnes = list(data[0].keys())
                writer.writerow(colonnes)
                for row in data:
                    writer.writerow([row.get(col, "") for col in colonnes])
            elif isinstance(data[0], list):
                for row in data:
                    writer.writerow(row)
            else:
                return "Erreur: format non supporte. Utilisez une liste de dicts ou de listes."

        nb_lignes = len(data)
        return f"Fichier CSV ecrit: '{Path(chemin).resolve()}' ({nb_lignes} lignes)"
    except json.JSONDecodeError as e:
        return f"Erreur JSON: {str(e)}"
    except Exception as e:
        return f"Erreur: {str(e)}"


def info_excel(chemin: str) -> str:
    """
    Retourne les informations sur un fichier Excel.

    Args:
        chemin: Chemin du fichier Excel

    Returns:
        Info: noms des feuilles, nombre de lignes/colonnes par feuille.
    """
    try:
        import openpyxl
    except ImportError:
        return "Erreur: openpyxl non installe. pip install openpyxl"

    try:
        path = Path(chemin)
        if not path.exists():
            return f"Erreur: le fichier '{chemin}' n'existe pas"

        wb = openpyxl.load_workbook(chemin, read_only=True)
        feuilles = []

        for name in wb.sheetnames:
            ws = wb[name]
            # Lire la premiere ligne pour les noms de colonnes
            first_row = []
            try:
                for row in ws.iter_rows(max_row=1, values_only=True):
                    first_row = [str(c) if c is not None else "" for c in row]
                    break
            except Exception:
                pass

            feuilles.append({
                "nom": name,
                "lignes": ws.max_row or 0,
                "colonnes": ws.max_column or 0,
                "noms_colonnes": first_row,
            })

        wb.close()

        return json.dumps(
            {
                "fichier": str(path.resolve()),
                "nombre_feuilles": len(feuilles),
                "feuilles": feuilles,
            },
            ensure_ascii=False,
            indent=2,
        )
    except Exception as e:
        return f"Erreur: {str(e)}"


def register_tools(mcp):
    """Enregistre les outils Excel/CSV sur l'instance MCP."""
    mcp.add_tool(lire_excel)
    mcp.add_tool(ecrire_excel)
    mcp.add_tool(lire_csv)
    mcp.add_tool(ecrire_csv)
    mcp.add_tool(info_excel)
